"""
XLSX Tool - Microsoft Excel spreadsheet processing.

Provides capabilities for:
- Reading Excel files
- Extracting sheet data
- Cell range access
- Basic data analysis
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from tools.base import Tool, ToolParameter, ToolResult, register_tool

logger = logging.getLogger("titan.tools.documents.xlsx")


class XLSXTool(Tool):
    """
    Microsoft Excel spreadsheet tool.

    Allows agents to read and manipulate XLSX files.
    """

    @property
    def name(self) -> str:
        return "xlsx"

    @property
    def description(self) -> str:
        return (
            "Microsoft Excel (XLSX) spreadsheet tool. "
            "Actions: 'sheets' (list sheets), 'read' (read sheet data), "
            "'cell' (read specific cell/range), 'summary' (get data summary)."
        )

    @property
    def parameters(self) -> list[ToolParameter]:
        return [
            ToolParameter(
                name="action",
                type="string",
                description="Action: 'sheets', 'read', 'cell', 'summary'",
                required=True,
                enum=["sheets", "read", "cell", "summary"],
            ),
            ToolParameter(
                name="path",
                type="string",
                description="Path to XLSX file",
                required=True,
            ),
            ToolParameter(
                name="sheet",
                type="string",
                description="Sheet name (default: active sheet)",
                required=False,
            ),
            ToolParameter(
                name="range",
                type="string",
                description="Cell range (e.g., 'A1:D10') for 'cell' action",
                required=False,
            ),
            ToolParameter(
                name="max_rows",
                type="integer",
                description="Maximum rows to return (default: 100)",
                required=False,
            ),
        ]

    async def execute(self, **kwargs: Any) -> ToolResult:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return ToolResult(
                success=False,
                output=None,
                error="openpyxl is required. Install with: pip install 'agentic-titan[documents]'",
            )

        action = kwargs.get("action", "")
        path = kwargs.get("path", "")
        sheet_name = kwargs.get("sheet")
        max_rows = kwargs.get("max_rows", 100)

        if not path:
            return ToolResult(
                success=False,
                output=None,
                error="'path' parameter is required",
            )

        try:
            wb = load_workbook(path, read_only=True, data_only=True)

            if action == "sheets":
                return ToolResult(
                    success=True,
                    output={
                        "sheets": wb.sheetnames,
                        "count": len(wb.sheetnames),
                        "active": wb.active.title if wb.active else None,
                    },
                )

            # Get the requested sheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return ToolResult(
                        success=False,
                        output=None,
                        error=f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}",
                    )
                ws = wb[sheet_name]
            else:
                ws = wb.active

            if action == "read":
                rows_data = []
                row_count = 0

                for row in ws.iter_rows(values_only=True):
                    if row_count >= max_rows:
                        break
                    # Convert to list and handle None values
                    row_values = [str(cell) if cell is not None else "" for cell in row]
                    # Skip completely empty rows
                    if any(v.strip() for v in row_values):
                        rows_data.append(row_values)
                        row_count += 1

                return ToolResult(
                    success=True,
                    output={
                        "sheet": ws.title,
                        "rows": len(rows_data),
                        "columns": ws.max_column,
                        "data": rows_data,
                        "truncated": ws.max_row > max_rows if ws.max_row else False,
                    },
                )

            elif action == "cell":
                cell_range = kwargs.get("range", "")
                if not cell_range:
                    return ToolResult(
                        success=False,
                        output=None,
                        error="'range' parameter required for cell action (e.g., 'A1:D10')",
                    )

                try:
                    cells = ws[cell_range]
                    # Handle single cell vs range
                    if hasattr(cells, "value"):
                        # Single cell
                        return ToolResult(
                            success=True,
                            output={
                                "range": cell_range,
                                "value": str(cells.value) if cells.value else "",
                            },
                        )
                    else:
                        # Range of cells
                        data = []
                        for row in cells:
                            row_data = [str(cell.value) if cell.value else "" for cell in row]
                            data.append(row_data)

                        return ToolResult(
                            success=True,
                            output={
                                "range": cell_range,
                                "data": data,
                            },
                        )
                except Exception as e:
                    return ToolResult(
                        success=False,
                        output=None,
                        error=f"Invalid range '{cell_range}': {e}",
                    )

            elif action == "summary":
                # Get basic stats about the sheet
                total_rows = ws.max_row or 0
                total_cols = ws.max_column or 0

                # Get column headers (first row)
                headers = []
                if total_rows > 0:
                    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                    headers = [str(h) if h else f"Col{i+1}" for i, h in enumerate(first_row)]

                # Sample some data types
                col_types: dict[int, set] = {}
                sample_rows = min(10, total_rows)

                for row in ws.iter_rows(min_row=2, max_row=sample_rows + 1, values_only=True):
                    for i, cell in enumerate(row):
                        if i not in col_types:
                            col_types[i] = set()
                        if cell is not None:
                            col_types[i].add(type(cell).__name__)

                column_info = []
                for i, header in enumerate(headers):
                    types = list(col_types.get(i, set()))
                    column_info.append({
                        "header": header,
                        "index": i,
                        "types": types,
                    })

                return ToolResult(
                    success=True,
                    output={
                        "sheet": ws.title,
                        "total_rows": total_rows,
                        "total_columns": total_cols,
                        "headers": headers,
                        "column_info": column_info[:20],  # Limit columns
                    },
                )

            else:
                return ToolResult(
                    success=False,
                    output=None,
                    error=f"Unknown action: {action}",
                )

        except FileNotFoundError:
            return ToolResult(
                success=False,
                output=None,
                error=f"File not found: {path}",
            )
        except Exception as e:
            logger.exception(f"XLSX tool error: {e}")
            return ToolResult(
                success=False,
                output=None,
                error=f"Failed to process XLSX: {e}",
            )


# Register the tool
xlsx_tool = XLSXTool()
register_tool(xlsx_tool)
